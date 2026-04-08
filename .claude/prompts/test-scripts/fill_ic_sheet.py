#!/usr/bin/env python3
"""
IC Sheet Auto-Filler
====================
Tự động sinh nội dung IC sheet trong file Excel Unit Test Report
từ một file C# xUnit test.

Usage
-----
  python fill_ic_sheet.py \\
      --excel "OISS_Report5_Unit Test_v0.9.xlsx" \\
      --sheet IC-20 \\
      --cs-file "intervu-be/Intervu.API.Test/ApiTests/QuestionController/ContributeToQuestionBank.cs" \\
      --function-name "Contribute to Question Bank" \\
      --created-by "YourName" \\
      --executed-by "YourName"

Options
-------
  --excel           Path to Excel file (required)
  --sheet           Sheet name to fill, e.g. IC-20 (required)
  --cs-file         Path to the .cs test file (required)
  --function-name   Function name to write in the sheet header (required)
  --created-by      Name of the person who created the test (default: "")
  --executed-by     Name of the person who ran the test (default: "")
  --date            Execution date YYYY-MM-DD (default: today)
  --lines-of-code   Lines of code for the function (default: 0)
  --preview         Print a table preview instead of modifying the file
  --force           Overwrite even if the sheet already has content
"""

import argparse
import sys
import os
from datetime import date
from pathlib import Path

# ── Ensure scripts/ is on sys.path ───────────────────────────────────────────
_SCRIPT_DIR = Path(__file__).parent
sys.path.insert(0, str(_SCRIPT_DIR))


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description='Auto-fill IC sheet in Unit Test Report Excel from a C# test file.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument('--excel',         required=True,  help='Path to the Excel file')
    p.add_argument('--sheet',         required=True,  help='Sheet name to fill (e.g. IC-20)')
    p.add_argument('--cs-file',       required=True,  help='Path to .cs unit test file', dest='cs_file')
    p.add_argument('--function-name', required=True,  help='Function name for sheet header', dest='function_name')
    p.add_argument('--created-by',    default='',     help='Creator name', dest='created_by')
    p.add_argument('--executed-by',   default='',     help='Executor name', dest='executed_by')
    p.add_argument('--date',          default=str(date.today()), help='Execution date YYYY-MM-DD')
    p.add_argument('--lines-of-code', default=0,      type=int, dest='lines_of_code')
    p.add_argument('--preview',       action='store_true', help='Print preview, do not write to file')
    p.add_argument('--force',         action='store_true', help='Overwrite even if sheet already has content')
    return p


def print_header(args):
    print()
    print('=' * 65)
    print('  IC Sheet Auto-Filler')
    print('=' * 65)
    print(f'  Excel        : {args.excel}')
    print(f'  Sheet        : {args.sheet}')
    print(f'  C# File      : {args.cs_file}')
    print(f'  Function     : {args.function_name}')
    print(f'  Date         : {args.date}')
    print(f'  Preview only : {args.preview}')
    print('=' * 65)
    print()


def print_preview(content):
    from ic_generator import IcSheetContent
    COL_WIDTHS = (4, 14, 26, 42, 30)
    header = f"{'Row':>4}  {'Section':<14}{'Sub-label':<26}{'Value':<42}{'Marks'}"
    print()
    print('[PREVIEW] Content to be written:')
    print('-' * 100)
    print(header)
    print('-' * 100)
    for i, row in enumerate(content.rows):
        marks_str = ', '.join(
            f'UTC{k:02d}={v}' for k, v in sorted(row.marks.items())
        )
        val_preview = (row.value or '')[:40]
        print(
            f'{10 + i:>4}  '
            f'{row.section_label or "":<14}'
            f'{row.sub_label or "":<26}'
            f'{val_preview:<42}'
            f'{marks_str}'
        )
    print('-' * 100)
    print(f'Total rows : {len(content.rows)}')
    print(f'Test cases : {content.n_cases}')
    print()


def force_clear_sheet(wb, sheet_name: str):
    """Remove O marks from a sheet so it looks 'unfilled' for re-processing."""
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=10):
        for cell in row:
            if cell.value == 'O':
                cell.value = None


def main():
    parser = build_parser()
    args = parser.parse_args()

    # ── Validate paths ────────────────────────────────────────────────────────
    if not os.path.isfile(args.excel):
        print(f'[ERROR] Excel file not found: {args.excel}')
        sys.exit(1)

    if not os.path.isfile(args.cs_file):
        print(f'[ERROR] C# test file not found: {args.cs_file}')
        sys.exit(1)

    print_header(args)

    # ── Step 1: Parse C# ─────────────────────────────────────────────────────
    from cs_parser import CSharpTestParser

    print('[1/3] Parsing C# test file ...')
    test_cases = CSharpTestParser().parse_file(args.cs_file)

    if not test_cases:
        print('  [WARN]  No [Fact] test methods found (all may be [Fact(Skip=...)]).')
        print('  Nothing to do.')
        sys.exit(0)

    print(f'  Found {len(test_cases)} test case(s):')
    for tc in test_cases:
        auth_info = f'[{tc.auth_role}]' if tc.requires_auth else '[Public]'
        print(
            f'    UTCID{tc.index:02d}  {tc.method_name}\n'
            f'            {tc.http_method} {tc.endpoint} '
            f'-> {tc.expected_status_text}  type={tc.test_type}  {auth_info}'
        )
        if tc.payload_fields:
            print(f'            payload: { {k: v for k, v in tc.payload_fields} }')
        if tc.query_params:
            print(f'            query  : {tc.query_params}')
    print()

    # ── Step 2: Generate content ──────────────────────────────────────────────
    from ic_generator import IcContentGenerator

    print('[2/3] Generating IC sheet content ...')
    content = IcContentGenerator().generate(
        test_cases=test_cases,
        sheet_name=args.sheet,
        function_name=args.function_name,
        created_by=args.created_by,
        executed_by=args.executed_by,
        executed_date=args.date,
        lines_of_code=args.lines_of_code,
    )
    print(f'  Generated {len(content.rows)} row(s).')
    print()

    # ── Preview only ──────────────────────────────────────────────────────────
    if args.preview:
        print_preview(content)
        return

    # ── Step 3: Write Excel ───────────────────────────────────────────────────
    import openpyxl
    from ic_filler import IcSheetFiller

    print('[3/3] Writing to Excel ...')
    wb = openpyxl.load_workbook(args.excel)

    if args.force and args.sheet in wb.sheetnames:
        print(f'  --force: clearing existing O marks in "{args.sheet}" ...')
        force_clear_sheet(wb, args.sheet)

    filler = IcSheetFiller()
    filled = filler.fill(wb, content)

    if filled:
        wb.save(args.excel)
        print()
        print(f'[OK]  Saved: {args.excel}')
    else:
        print()
        print('[SKIP]   Nothing written. Use --force to overwrite an already-filled sheet.')


if __name__ == '__main__':
    main()
