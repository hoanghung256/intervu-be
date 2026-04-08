# ic_filler.py
# Write IcSheetContent into an openpyxl Workbook.
#
# DESIGN PRINCIPLE:
#   When the target sheet already EXISTS (template prepared by user):
#     - NEVER delete/recreate the sheet
#     - NEVER call _apply_styles
#     - Only write .value into:
#         • a few specific metadata cells (IC code, function name, etc.)
#         • col D  (content values)
#         • col F+ (UTCID O-marks)
#     - Col A, B, C, E and all styles are left completely untouched
#
#   When the target sheet does NOT exist yet:
#     - Create it fresh and also apply basic styles.

import re
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ic_generator import IcSheetContent, IcRow

# ── Fixed row positions (1-based) ────────────────────────────────────────────
ROW_FUNC_CODE  = 2
ROW_CREATED_BY = 3
ROW_LINES_CODE = 4
ROW_STATS_HDR  = 6
ROW_STATS      = 7
ROW_UTCID      = 9
ROW_DATA_START = 10

# ── Fixed column positions (1-based) ────────────────────────────────────────
COL_A = 1   # Section label
COL_B = 2   # Sub-label
COL_C = 3   # (merged / IC code value)
COL_D = 4   # Content value
COL_E = 5   # Empty spacer — DO NOT TOUCH
COL_F = 6   # First UTCID column (default)

# Metadata cell positions
COL_IC_CODE_VAL   = 3   # C: "IC-20"
COL_FUNC_NAME_LBL = 6   # F: "Function Name"  label
COL_FUNC_NAME_VAL = 12  # L: actual function name
COL_EXEC_LBL      = 6   # F: "Executed By"    label
COL_EXEC_VAL      = 12  # L: executed-by name
COL_TOTAL_TC      = 15  # O: total test cases

# ── Section label patterns (col B) ───────────────────────────────────────────
_SECTIONS = {
    'precondition': re.compile(r'precondition', re.I),
    'input':        re.compile(r'^\s*input\s*$', re.I),
    'http_method':  re.compile(r'http\s*method', re.I),
    'api_endpoint': re.compile(r'api\s*endpoint', re.I),
    'return_':      re.compile(r'^\s*return\s*$', re.I),
    'exception':    re.compile(r'^\s*exception', re.I),
    'log_message':  re.compile(r'log\s*message', re.I),
    'type_':        re.compile(r'^\s*type\s*[\(\[]\s*n', re.I),
    'passed_failed':re.compile(r'passed.*/?\s*failed', re.I),
    'executed_date':re.compile(r'executed\s*date', re.I),
    'defect_id':    re.compile(r'defect\s*id', re.I),
}


class IcSheetFiller:

    # ── Public API ────────────────────────────────────────────────────────────

    def fill(self, wb: openpyxl.Workbook, content: IcSheetContent) -> bool:
        """
        Fill the target sheet.
        Returns True if filled (caller should save), False if skipped.
        """
        name = content.sheet_name

        if name in wb.sheetnames:
            ws = wb[name]
            if self._is_filled(ws):
                print(f'  -> Skip: sheet "{name}" already has O marks.')
                return False
            print(f'  Template mode: writing values only (style preserved).')
            self._fill_into_template(ws, content)
        else:
            ws = wb.create_sheet(name)
            print(f'  Fresh mode: creating new sheet with styles.')
            self._fill_fresh(ws, content)

        n_rows = len(content.rows)
        print(f'  [OK] Sheet "{name}" filled — {content.n_cases} test case(s), {n_rows} data rows.')
        return True

    # ── Filled detection ──────────────────────────────────────────────────────

    def _is_filled(self, ws) -> bool:
        """True if any 'O' exists in UTCID columns (col F+) at rows 10+."""
        for row in ws.iter_rows(min_row=ROW_DATA_START, max_col=60, values_only=True):
            if row and any(v == 'O' for v in row[COL_F - 1:]):
                return True
        return False

    # ═════════════════════════════════════════════════════════════════════════
    # TEMPLATE MODE  —  existing sheet, preserve all styles
    # ═════════════════════════════════════════════════════════════════════════

    def _fill_into_template(self, ws, content: IcSheetContent):
        """
        Write only values:
          • A handful of metadata cells (IC code, function name, etc.)
          • col D  — content values
          • col F+ — UTCID marks / N/A/B / P / date
        Everything else (col A, B, C, E, all styles) is left untouched.
        """
        utcid_col = self._find_utcid_start_col(ws)

        # 1. Metadata cells only
        self._write_meta_values(ws, content)

        # 2. UTCID headers (row 9)
        for i in range(1, content.n_cases + 1):
            ws.cell(ROW_UTCID, utcid_col + i - 1).value = f'UTCID{i:02d}'

        # 3. Find section header rows in the template
        sections = self._scan_sections(ws)

        # 4. Clear ONLY col D and UTCID-range cells in the data area
        #    (so we can write fresh; structure labels in A/B/E stay)
        self._clear_data_cells(ws, utcid_col, content.n_cases)

        # 5. Write each section
        self._write_all_sections(ws, sections, utcid_col, content)

    # ── Helpers for template mode ─────────────────────────────────────────────

    def _find_utcid_start_col(self, ws) -> int:
        """Find the column of UTCID01 in row 9 (default COL_F if not found)."""
        for cell in ws[ROW_UTCID]:
            if cell.value and str(cell.value).strip().upper().startswith('UTCID'):
                return cell.column
        return COL_F

    def _scan_sections(self, ws) -> dict:
        """
        Scan col B (and col A for 'Confirm'/'Result') from ROW_DATA_START downwards.
        Returns {section_key: row_number}.
        """
        found = {}
        for row in ws.iter_rows(min_row=ROW_DATA_START, max_row=200):
            b_cell = next((c for c in row if c.column == COL_B), None)
            if b_cell and b_cell.value:
                val = str(b_cell.value).strip()
                for key, pattern in _SECTIONS.items():
                    if key not in found and pattern.search(val):
                        found[key] = b_cell.row
                        break
        return found

    def _clear_data_cells(self, ws, utcid_col: int, n_cases: int):
        """
        Clear ONLY values in col D and ALL UTCID columns (from n_cases+1 onwards).
        This removes redundant "O" marks and headers if the template was too large.
        """
        # 1. Clear UTCID headers (row 9) from the first UTCID col to the very end of sheet
        for c in range(utcid_col, ws.max_column + 1):
            _set(ws, ROW_UTCID, c, None)

        # 2. Clear data area (rows 10+)
        # We clear COL_D and ALL columns from utcid_col to the right
        for r in range(ROW_DATA_START, ws.max_row + 1):
            # Clear content value
            _set(ws, r, COL_D, None)
            # Clear all marker columns
            for c in range(utcid_col, ws.max_column + 1):
                _set(ws, r, c, None)

    def _write_meta_values(self, ws, content: IcSheetContent):
        """Write only the value cells in the header area (rows 2-7)."""
        _set(ws, ROW_FUNC_CODE,  COL_IC_CODE_VAL, content.sheet_name)
        _set(ws, ROW_FUNC_CODE,  COL_FUNC_NAME_VAL, content.function_name)
        _set(ws, ROW_CREATED_BY, COL_IC_CODE_VAL, content.created_by)
        _set(ws, ROW_CREATED_BY, COL_EXEC_VAL,    content.executed_by)
        _set(ws, ROW_STATS, COL_A,        content.n_cases)   # passed count
        _set(ws, ROW_STATS, 3,            0)                  # failed
        _set(ws, ROW_STATS, COL_F,        0)                  # untested
        _set(ws, ROW_STATS, COL_TOTAL_TC, content.n_cases)    # total

    def _write_all_sections(self, ws, sections: dict, utcid_col: int, content: IcSheetContent):
        """
        For each section found in the template, write generated rows into the
        empty D cells and UTCID marker cells that follow the section header.
        Section boundaries are determined by the next section header's row.
        """
        # Build an ordered list of (section_key, header_row, next_header_row)
        # so we know how many rows each section has available.
        order = [
            'precondition', 'input', 'http_method', 'api_endpoint',
            'return_', 'exception', 'log_message',
            'type_', 'passed_failed', 'executed_date', 'defect_id',
        ]
        # All rows ordered by position in the sheet
        rows_by_section = sorted(
            [(k, v) for k, v in sections.items()],
            key=lambda x: x[1]
        )
        boundaries = {}
        for i, (key, row) in enumerate(rows_by_section):
            next_row = rows_by_section[i + 1][1] if i + 1 < len(rows_by_section) else 300
            boundaries[key] = (row, row + 1, next_row - 1)  # (header, first_data, last_data)

        # Separate IcRows by section
        prec_rows, inp_rows, http_rows, ep_rows = [], [], [], []
        ret_rows, log_rows, exc_rows = [], [], []

        section_cursor = None
        for ic_row in content.rows:
            # Detect which logical section this row belongs to
            sl = ic_row.section_label.lower() if ic_row.section_label else ''
            sb = ic_row.sub_label.lower()   if ic_row.sub_label   else ''

            if 'precondition' in sb:
                section_cursor = 'precondition'; continue
            elif sb == 'input':
                section_cursor = 'input'; continue
            elif 'http method' in sb:
                section_cursor = 'http_method'; continue
            elif 'api endpoint' in sb:
                section_cursor = 'api_endpoint'; continue
            elif sb == 'return':
                section_cursor = 'return_'; continue
            elif 'exception' in sb:
                section_cursor = 'exception'; continue
            elif 'log message' in sb:
                section_cursor = 'log_message'; continue
            elif 'type' in sb:
                section_cursor = 'type_'; continue
            elif 'passed' in sb:
                section_cursor = 'passed_failed'; continue
            elif 'executed' in sb:
                section_cursor = 'executed_date'; continue
            elif 'defect' in sb:
                section_cursor = 'defect_id'; continue

            if ic_row.is_header:
                continue

            # Route this data row to the correct bucket
            if section_cursor == 'precondition':   prec_rows.append(ic_row)
            elif section_cursor == 'input':         inp_rows.append(ic_row)
            elif section_cursor == 'http_method':   http_rows.append(ic_row)
            elif section_cursor == 'api_endpoint':  ep_rows.append(ic_row)
            elif section_cursor == 'return_':       ret_rows.append(ic_row)
            elif section_cursor == 'exception':     exc_rows.append(ic_row)
            elif section_cursor == 'log_message':   log_rows.append(ic_row)
            # Result rows handled separately below

        # ── Write each section into the template ──────────────────────────────
        def write_section(key, data_rows):
            if key not in boundaries:
                return
            _, first, last = boundaries[key]
            cur = first
            for ic_row in data_rows:
                if cur > last:
                    print(f'  [WARN] More rows than template space in section "{key}" — some rows skipped.')
                    break
                if ic_row.value:
                    _set(ws, cur, COL_D, ic_row.value)
                for utcid_idx, mark in ic_row.marks.items():
                    val = content.executed_date if mark == '__DATE__' else mark
                    _set(ws, cur, utcid_col + utcid_idx - 1, val)
                cur += 1

        write_section('precondition', prec_rows)
        write_section('input',        inp_rows)
        write_section('http_method',  http_rows)
        write_section('api_endpoint', ep_rows)
        write_section('return_',      ret_rows)
        write_section('exception',    exc_rows)
        write_section('log_message',  log_rows)

        # ── Result section rows (each sub-label has its own row) ──────────────
        for key, sub in [
            ('type_',        'type_'),
            ('passed_failed','passed_failed'),
            ('executed_date','executed_date'),
            ('defect_id',    'defect_id'),
        ]:
            if key not in sections:
                continue
            r = sections[key]
            # Find the matching IcRow
            for ic_row in content.rows:
                sb = (ic_row.sub_label or '').lower()
                is_match = (
                    (key == 'type_'         and 'type' in sb) or
                    (key == 'passed_failed' and 'passed' in sb) or
                    (key == 'executed_date' and 'executed' in sb) or
                    (key == 'defect_id'     and 'defect' in sb)
                )
                if is_match:
                    for utcid_idx, mark in ic_row.marks.items():
                        val = content.executed_date if mark == '__DATE__' else mark
                        _set(ws, r, utcid_col + utcid_idx - 1, val)
                    break

    # ═════════════════════════════════════════════════════════════════════════
    # FRESH MODE  —  new sheet, create structure + apply styles
    # ═════════════════════════════════════════════════════════════════════════

    def _fill_fresh(self, ws, content: IcSheetContent):
        """Write a brand-new sheet from scratch, then apply basic styles."""
        utcid_col = COL_F

        # Metadata labels + values
        _set(ws, ROW_FUNC_CODE,  COL_A,           'Function Code')
        _set(ws, ROW_FUNC_CODE,  COL_IC_CODE_VAL, content.sheet_name)
        _set(ws, ROW_FUNC_CODE,  COL_FUNC_NAME_LBL,'Function Name')
        _set(ws, ROW_FUNC_CODE,  COL_FUNC_NAME_VAL, content.function_name)
        _set(ws, ROW_CREATED_BY, COL_A,           'Created By')
        _set(ws, ROW_CREATED_BY, COL_IC_CODE_VAL, content.created_by)
        _set(ws, ROW_CREATED_BY, COL_EXEC_LBL,    'Executed By')
        _set(ws, ROW_CREATED_BY, COL_EXEC_VAL,    content.executed_by)
        _set(ws, ROW_LINES_CODE, COL_A,           'Lines  of code')
        if content.lines_of_code:
            _set(ws, ROW_LINES_CODE, COL_IC_CODE_VAL, content.lines_of_code)
        _set(ws, ROW_LINES_CODE, COL_EXEC_LBL,    'Lack of test cases')
        _set(ws, ROW_STATS_HDR,  COL_A,           'Passed')
        _set(ws, ROW_STATS_HDR,  3,               'Failed')
        _set(ws, ROW_STATS_HDR,  COL_F,           'Untested')
        _set(ws, ROW_STATS_HDR,  12,              'N/A/B')
        _set(ws, ROW_STATS_HDR,  COL_TOTAL_TC,    'Total Test Cases')
        _set(ws, ROW_STATS,      COL_A,           content.n_cases)
        _set(ws, ROW_STATS,      3,               0)
        _set(ws, ROW_STATS,      COL_F,           0)
        _set(ws, ROW_STATS,      COL_TOTAL_TC,    content.n_cases)

        # UTCID headers
        for i in range(1, content.n_cases + 1):
            _set(ws, ROW_UTCID, utcid_col + i - 1, f'UTCID{i:02d}')

        # Data rows (sequential)
        cur = ROW_DATA_START
        for ic_row in content.rows:
            if ic_row.section_label:
                _set(ws, cur, COL_A, ic_row.section_label)
            if ic_row.sub_label:
                _set(ws, cur, COL_B, ic_row.sub_label)
            if ic_row.value and not ic_row.is_header:
                _set(ws, cur, COL_D, ic_row.value)
            for utcid_idx, mark in ic_row.marks.items():
                val = content.executed_date if mark == '__DATE__' else mark
                _set(ws, cur, utcid_col + utcid_idx - 1, val)
            cur += 1

        last_row = cur - 1
        self._apply_styles(ws, content, last_row, utcid_col)

    # ── Styles (only used for fresh sheets) ──────────────────────────────────

    def _apply_styles(self, ws, content: IcSheetContent, last_row: int, utcid_col: int):
        n = content.n_cases
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left   = Alignment(vertical='center', wrap_text=True)
        bold   = Font(bold=True)

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 6
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 4
        for i in range(1, n + 1):
            ws.column_dimensions[get_column_letter(utcid_col + i - 1)].width = 14

        for r in range(ROW_DATA_START, last_row + 1):
            for c in range(1, utcid_col + n):
                cell = ws.cell(r, c)
                if c >= utcid_col:
                    cell.alignment = center
                else:
                    cell.alignment = left
                if c in (COL_A, COL_B) and cell.value:
                    cell.font = bold

        for i in range(1, n + 1):
            cell = ws.cell(ROW_UTCID, utcid_col + i - 1)
            cell.alignment = center
            cell.font = bold


# ── Tiny helper ───────────────────────────────────────────────────────────────

def _set(ws, row: int, col: int, value):
    """Write a value to a cell, ignoring MergedCell slaves gracefully."""
    try:
        ws.cell(row, col).value = value
    except AttributeError:
        pass  # merged cell slave — skip silently
